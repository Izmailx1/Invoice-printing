import datetime
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import PySimpleGUI as sg
import os


def generate_invoice(item_list, item_count, container='x', name='Имя'):
    invoice = Workbook()
    invoice_sheet = invoice.active
    invoice_sheet.row_dimensions[2].height = 30
    invoice_sheet.column_dimensions['A'].width = 4
    invoice_sheet.column_dimensions['B'].width = 50
    invoice_sheet.column_dimensions['C'].width = 8
    invoice_sheet.column_dimensions['D'].width = 5.5
    invoice_sheet.column_dimensions['E'].width = 4
    invoice_sheet.column_dimensions['F'].width = 7
    invoice_sheet.column_dimensions['G'].width = 8
    invoice_sheet.merge_cells('A1:G1')
    date = datetime.date.today()
    invoice_sheet.cell(1, 1).value = f'Накладная от {date} Контейнер {container}'
    invoice_sheet.cell(2, 2).value = f'Наименование'
    invoice_sheet.cell(2, 3).value = f'Произв.'
    invoice_sheet.cell(2, 4).value = f'Ед.'
    invoice_sheet.cell(2, 5).value = f'Кол-во'
    invoice_sheet.cell(2, 6).value = f'Цена'
    invoice_sheet.cell(2, 7).value = f'Сумма'

    for i in range(len(item_count)):
        invoice_sheet.cell(invoice_sheet.max_row + 1, 1).value = '=ROW()-ROW($A$2)'
        invoice_sheet.cell(invoice_sheet.max_row, 2).value = item_count[i][0]
        invoice_sheet.cell(invoice_sheet.max_row, 3).value = item_list[item_count[i][0]][0]
        invoice_sheet.cell(invoice_sheet.max_row, 4).value = item_list[item_count[i][0]][1]
        invoice_sheet.cell(invoice_sheet.max_row, 5).value = item_count[i][1]
        invoice_sheet.cell(invoice_sheet.max_row, 6).value = item_count[i][2]
        invoice_sheet.cell(invoice_sheet.max_row, 7).value = f'=E{invoice_sheet.max_row}*F{invoice_sheet.max_row}'

    invoice_sheet.cell(invoice_sheet.max_row + 1, 5).value = 'Итого:'
    invoice_sheet.merge_cells(start_row=invoice_sheet.max_row, start_column=5,
                              end_row=invoice_sheet.max_row, end_column=6)
    invoice_sheet.cell(invoice_sheet.max_row, 7).value = f'=SUM(G3:G{invoice_sheet.max_row - 1})'

    for i in range(1, 8):
        invoice_sheet.cell(2, i).alignment = Alignment(horizontal='center', vertical="center")
        invoice_sheet.cell(2, i).fill = PatternFill('solid', fgColor='AFEEEE')

    for row in range(1, invoice_sheet.max_row + 1):
        for col in range(1, 8):
            invoice_sheet.cell(row, col).font = font
            invoice_sheet.cell(row, col).border = border
            invoice_sheet.cell(row, col).alignment = Alignment(wrap_text=True)

    invoice.save(f'{desktop}\{date}_{container}_{name}.xlsx')


desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
thins = Side(border_style="thin", color="000000")
font = Font(name='Bahnschrift SemiBold', size=10)
border = Border(top=thins, bottom=thins, left=thins, right=thins)
price = xl.load_workbook('D:\WORK\Avtoviraz\Прайс.xlsx', data_only=True)
price_sheet = price['Лист1']
items = {}

for i in range(1, price_sheet.max_row + 1):
    try:
        if isinstance(int(price_sheet.cell(i, 1).value), int):
            temp = []
            for j in range(2, 8):
                temp.append(price_sheet.cell(i, j).value)
            items[temp[0]] = temp[1:]
    except:
        pass

keys = list(items)
print_items = []
res_s = ''
headings = ['            Наименование                  ', 'Количество', 'Цена']
background_text_color = '#FFEBCD'
text_text_color = '#808080'
background_input_color = '#FFFFE0'
background_button_color = '#3CB371'

layout = [
    [sg.Text('Имя клиента', size=(20, 1), font='Bahnschrift', justification='left', background_color=background_text_color, text_color=text_text_color)],
    [sg.InputText(key='_client_', size=(15, 10), background_color=background_input_color)],
    [sg.Text('Номер контейнера (если есть)', size=(30, 1), font='Bahnschrift', justification='left', background_color=background_text_color, text_color=text_text_color)],
    [sg.InputText(key='_container_', size=(10, 10), background_color=background_input_color)],
    [sg.Text('Выбери наименование', size=(20, 1), font='Bahnschrift', justification='left', background_color=background_text_color, text_color=text_text_color)],
    [sg.Combo(keys, default_value=keys[0], key='_list_items_', enable_events=True, background_color=background_input_color)],
    [sg.Text('Введи количество', size=(20, 1), font='Bahnschrift', justification='left', background_color=background_text_color, text_color=text_text_color)],
    [sg.InputText(key='_quantity_', background_color=background_input_color)],
    [sg.Text('Введи цену', size=(20, 1), font='Bahnschrift', justification='left', background_color=background_text_color, text_color=text_text_color)],
    [sg.InputText(key='_price_', background_color=background_input_color)],
    [sg.Button('Добавить в накладную', font=('Bahnschrift SemiBold', 12), key='_add_', button_color=background_button_color)],
    [sg.Table(values=print_items, headings=headings, max_col_width=95, background_color=background_input_color,
                    size=(888, 80),
                    justification='center',
                    num_rows=20,
                    auto_size_columns=True,
                    key='-TABLE-',
                    expand_x=True,
                    expand_y=True,
                    row_height=15,
                    text_color='#000000',
                    header_background_color='#FAEBD7',
                    font=('Bahnschrift SemiBold', 11))],
    [sg.Button('Сбросить', font=('Bahnschrift SemiBold', 12), key='_reset_', button_color='#FF7F50'),
     sg.Button('Печатать накладную', font=('Bahnschrift SemiBold', 12), key='_print_', button_color=background_button_color)]
]

window = sg.Window('Накладная', layout, resizable=True, background_color='#FFEBCD',
                   font=('Bahnschrift SemiBold', 10), size=(1000, 700))

while True:
    event, values = window.read()
    if event in (None, 'Exit', 'Cancel'):
        break
    if event == '_add_':
        pos = []
        pos.append(values['_list_items_'])
        s = 'Наим.: '
        s += pos[-1]
        s += ' '
        pos.append(values['_quantity_'])
        s += 'Кол-во: '
        s += pos[-1]
        s += ' '
        pos.append(values['_price_'])
        s += 'Цена: '
        s += pos[-1]
        print_items.append(pos)
        window['_list_items_'].update('')
        window['_quantity_'].update('')
        window['_price_'].update('')
        res_s += s
        res_s += '\n'
        window['-TABLE-'].update(print_items)

    if event == '_reset_':
        window['-TABLE-'].update('')
        print_items = []
        res_s = ''

    if event == '_print_':
        generate_invoice(items, print_items, values['_container_'], values['_client_'])
        sg.popup('Накладная сохранена', res_s, background_color=background_input_color, text_color='#000000')

    if event == '_list_items_':
        window['_price_'].update(items[f"{values['_list_items_']}"][4])
